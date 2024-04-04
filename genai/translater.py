import pandas as pd
from pandas.io.formats import excel
import os
import shutil
from openpyxl import load_workbook
from transformers import AutoModelForSeq2SeqLM, AutoTokenizer

def get_xlsx_files(folder_path) -> list:
    xlsx_files = [os.path.join(folder_path, file) for file in os.listdir(folder_path) if file.lower().endswith('.xlsx')]
    return xlsx_files

def read_excel_to_dataframe(file_path: str, sheet_name: str = None, usecols: list = None) -> pd.DataFrame:
    try:
        df = pd.read_excel(file_path, sheet_name=sheet_name, usecols=usecols)
        return df
    except FileNotFoundError:
        print(f"Error: File '{file_path}' not found.")
        return None

def copy_excel_file(source_path, destination_path):
    try:
        shutil.copy(source_path, destination_path)
        return f"Excel file copied from '{source_path}' to '{destination_path}' successfully."
    except Exception as e:
        return f"Error: {str(e)}"
    
def duplicate_sheet(input_file_path, sheet_name, output_file_path, new_sheet_name):
    try:
        # Load the workbook
        wb = load_workbook(input_file_path)

        # Get the original sheet
        original_sheet = wb[sheet_name]

        # check if Translated_* version exists
        if new_sheet_name in wb.sheetnames:
            wb.remove(worksheet=wb[new_sheet_name])
        
        # Create a new sheet by copying the original
        new_sheet = wb.copy_worksheet(original_sheet)
        new_sheet.title = new_sheet_name
        
        # Save the modified workbook
        wb.save(output_file_path)
        print(f"Sheet '{sheet_name}' duplicated as '{new_sheet_name}' in '{output_file_path}'.")

    except Exception as e:
            print(f"Error: {e}")
        

tokenizer = AutoTokenizer.from_pretrained("facebook/nllb-200-distilled-600M", src_lang="hun_Latn")
model = AutoModelForSeq2SeqLM.from_pretrained("facebook/nllb-200-distilled-600M")

def translate(df: pd.DataFrame = None, column_name: str = None):
    translated_texts = []

    # working on singular text - in this case: column name of dataframe
    if df is None and column_name is not None:
        model_inputs = tokenizer(str(column_name), return_tensors="pt")
        gen_tokens = model.generate(**model_inputs, forced_bos_token_id=tokenizer.lang_code_to_id["eng_Latn"])
        translated_text = tokenizer.batch_decode(gen_tokens, skip_special_tokens=True)
        return translated_text
    # working on dataframe
    else:
        for text in df[column_name]:
            model_inputs = tokenizer(str(text), return_tensors="pt")
            gen_tokens = model.generate(**model_inputs, forced_bos_token_id=tokenizer.lang_code_to_id["eng_Latn"])
            translated_text = tokenizer.batch_decode(gen_tokens, skip_special_tokens=True)
            translated_texts.append(translated_text)

        df['translated_' + column_name] = translated_texts
        return df

def run_translater_on_folder(folder_path: str):
    # set dataframe header style
    excel.ExcelFormatter.header_style = None
    # get all excel files inside folder
    excel_files = get_xlsx_files(folder_path=folder_path)

    for excel_file in excel_files:
        # get all the sheets inside the current excel file
        df_dict = pd.read_excel(excel_file, sheet_name=None)

        sheet_names = list(df_dict.keys())

        # go through all sheets
        for sheet in sheet_names:
            # only check original sheets, not Translated_* ones
            if 'Translated_' not in str(sheet):
                print('Working on', str(excel_file), ', sheet:', str(sheet))

                # copy sheet with content, formatting and adding Translated_ prefix
                new_sheet_name = 'Translated_' + str(sheet)
                duplicate_sheet(
                    input_file_path=excel_file, 
                    sheet_name=str(sheet), 
                    output_file_path=excel_file,
                    new_sheet_name=new_sheet_name
                    )

                df = df_dict[sheet]
                
                # get column name
                column_name = df.columns[0]

                # inference model
                translated_df = translate(df=df, column_name=column_name)

                # rename original column with translated version
                original_translated_column_name = 'translated_' + column_name
                new_translated_column_name = translate(column_name=column_name)[0]
                translated_df = translated_df.rename(columns={original_translated_column_name: new_translated_column_name})

                # drop original (not-translated) column
                translated_df = translated_df.drop([str(column_name)], axis=1)

                # extract string values from generated translation
                translated_df[new_translated_column_name] = translated_df[new_translated_column_name].str.get(0)

                # write result to new Translated_* sheet
                try:
                    with pd.ExcelWriter(excel_file, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
                        translated_df.to_excel(writer, sheet_name=new_sheet_name, index=False)
                        print(f"Data written to sheet '{new_sheet_name}' in '{excel_file}' successfully.")
                except Exception as e:
                    print(f"Error: {e}")
            
# run the whole thing
folder_path = "C:\\Users\\richa\\Source\\Repos\\learning\\genai\\"
run_translater_on_folder(folder_path=folder_path)
